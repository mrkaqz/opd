import io
import os
import shutil
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy import or_, select, func
from sqlalchemy.orm import Session, selectinload

from app.database import get_db, DB_PATH
from app.models import OpdVisit, OpdPatient, OpdPhone, OpdOwner
from app.schemas import (
    VisitCreate, VisitOut, VisitSummary, PaginatedVisits,
    PhoneSearchResult, PatientOut, PhoneOut, OwnerOut, ImportResult,
)
from app.services import importer as importer_svc

router = APIRouter(prefix="/api/visits", tags=["visits"])


def _load_opts():
    return [
        selectinload(OpdVisit.patients),
        selectinload(OpdVisit.phones),
        selectinload(OpdVisit.owners),
    ]


def _visit_to_summary(visit: OpdVisit) -> VisitSummary:
    owners    = ", ".join(o.owner_name for o in visit.owners)
    pets      = ", ".join(p.pet_name for p in visit.patients if p.pet_name)
    pet_types = ", ".join({p.pet_type for p in visit.patients if p.pet_type})
    phones    = ", ".join(p.phone for p in visit.phones)
    return VisitSummary(
        opd_number=visit.opd_number,
        owners=owners, pets=pets, pet_types=pet_types, phones=phones,
        has_file=bool(visit.web_url),
        file_name=visit.file_name, web_url=visit.web_url,
        patient_count=len(visit.patients),
    )


def _sort_expr(sort_by: str, sort_dir: str):
    """Return a SQLAlchemy order expression for the requested column."""
    desc = sort_dir.lower() == "desc"

    if sort_by == "owners":
        sub = (select(func.min(OpdOwner.owner_name))
               .where(OpdOwner.opd_number == OpdVisit.opd_number)
               .correlate(OpdVisit)
               .scalar_subquery())
    elif sort_by == "pets":
        sub = (select(func.min(OpdPatient.pet_name))
               .where(OpdPatient.opd_number == OpdVisit.opd_number)
               .correlate(OpdVisit)
               .scalar_subquery())
    elif sort_by == "pet_types":
        sub = (select(func.min(OpdPatient.pet_type))
               .where(OpdPatient.opd_number == OpdVisit.opd_number)
               .correlate(OpdVisit)
               .scalar_subquery())
    elif sort_by == "phones":
        sub = (select(func.min(OpdPhone.phone))
               .where(OpdPhone.opd_number == OpdVisit.opd_number)
               .correlate(OpdVisit)
               .scalar_subquery())
    elif sort_by == "has_file":
        col = OpdVisit.web_url
        return col.desc() if desc else col.asc()
    else:  # opd_number (default)
        col = OpdVisit.opd_number
        return col.desc() if desc else col.asc()

    return sub.desc() if desc else sub.asc()


@router.get("", response_model=PaginatedVisits)
def list_visits(
    q: str | None = Query(None),
    opd: int | None = Query(None),
    pet_type: str | None = Query(None),
    sort_by: str = Query("opd_number"),
    sort_dir: str = Query("desc"),
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    # ── Step 1: build a query that returns just OPD numbers ──────────────────
    # This keeps DISTINCT clean and allows the scalar-subquery ORDER BY to work.
    id_q = db.query(OpdVisit.opd_number)

    if opd is not None:
        id_q = id_q.filter(OpdVisit.opd_number == opd)

    # Only join when we actually need to filter on related tables
    if q or pet_type:
        id_q = (id_q
                .outerjoin(OpdPatient, OpdVisit.opd_number == OpdPatient.opd_number)
                .outerjoin(OpdPhone,   OpdVisit.opd_number == OpdPhone.opd_number)
                .outerjoin(OpdOwner,   OpdVisit.opd_number == OpdOwner.opd_number))

        if pet_type:
            id_q = id_q.filter(OpdPatient.pet_type.ilike(f"%{pet_type}%"))

        if q:
            like = f"%{q}%"
            try:
                opd_q = int(q)
                id_q = id_q.filter(or_(
                    OpdVisit.opd_number == opd_q,
                    OpdPhone.phone.ilike(like),
                    OpdOwner.owner_name.ilike(like),
                    OpdPatient.pet_name.ilike(like),
                ))
            except ValueError:
                id_q = id_q.filter(or_(
                    OpdPhone.phone.ilike(like),
                    OpdOwner.owner_name.ilike(like),
                    OpdPatient.pet_name.ilike(like),
                ))

        id_q = id_q.distinct()

    # ── Step 2: count total, then get the page of IDs in sort order ──────────
    total = id_q.count()
    ordered_ids = [
        row[0] for row in
        id_q.order_by(_sort_expr(sort_by, sort_dir))
            .offset((page - 1) * limit)
            .limit(limit)
            .all()
    ]

    if not ordered_ids:
        return PaginatedVisits(total=total, page=page, limit=limit, items=[])

    # ── Step 3: fetch full visit objects for those IDs ───────────────────────
    visit_map = {
        v.opd_number: v
        for v in db.query(OpdVisit)
                   .options(*_load_opts())
                   .filter(OpdVisit.opd_number.in_(ordered_ids))
                   .all()
    }
    # Restore the sort order from step 2
    visits = [visit_map[oid] for oid in ordered_ids if oid in visit_map]

    return PaginatedVisits(
        total=total, page=page, limit=limit,
        items=[_visit_to_summary(v) for v in visits],
    )


@router.get("/search-phone", response_model=list[PhoneSearchResult])
def search_by_phone(
    phone: str = Query(..., min_length=1),
    db: Session = Depends(get_db),
):
    matching = (
        db.query(OpdPhone.opd_number)
        .filter(OpdPhone.phone == phone.strip())
        .subquery()
    )
    visits = (
        db.query(OpdVisit)
        .options(*_load_opts())
        .filter(OpdVisit.opd_number.in_(matching))
        .order_by(OpdVisit.opd_number.desc())
        .all()
    )
    return [
        PhoneSearchResult(
            opd_number=v.opd_number,
            phones=[PhoneOut.model_validate(ph) for ph in v.phones],
            owners=[OwnerOut.model_validate(o) for o in v.owners],
            patients=[PatientOut.model_validate(p) for p in v.patients],
        )
        for v in visits
    ]


@router.post("", response_model=VisitOut, status_code=201)
def create_visit(body: VisitCreate, db: Session = Depends(get_db)):
    existing = db.get(OpdVisit, body.opd_number)
    if existing:
        raise HTTPException(status_code=409, detail=f"OPD {body.opd_number} already exists")

    visit = OpdVisit(opd_number=body.opd_number)
    db.add(visit)
    db.flush()

    db.add(OpdPhone(opd_number=body.opd_number, phone=body.first_phone))
    if body.first_owner and body.first_owner.strip():
        db.add(OpdOwner(opd_number=body.opd_number, owner_name=body.first_owner.strip()))
    db.add(OpdPatient(opd_number=body.opd_number, **body.first_patient.model_dump()))
    db.commit()
    db.refresh(visit)
    return visit


@router.get("/{opd_number}", response_model=VisitOut)
def get_visit(opd_number: int, db: Session = Depends(get_db)):
    visit = (
        db.query(OpdVisit)
        .options(*_load_opts())
        .filter(OpdVisit.opd_number == opd_number)
        .first()
    )
    if not visit:
        raise HTTPException(status_code=404, detail="OPD not found")
    return visit


@router.delete("/{opd_number}", status_code=204)
def delete_visit(opd_number: int, db: Session = Depends(get_db)):
    visit = db.get(OpdVisit, opd_number)
    if not visit:
        raise HTTPException(status_code=404, detail="OPD not found")
    db.delete(visit)
    db.commit()


# ── Admin: import ────────────────────────────────────────────────────────────
admin_router = APIRouter(prefix="/api/admin", tags=["admin"])


@admin_router.post("/import", response_model=ImportResult)
def trigger_import(
    file: UploadFile = File(..., description="Patient List .xlsx/.xlsm file"),
    db: Session = Depends(get_db),
):
    data = file.file.read()
    return importer_svc.run_import(db, file_bytes=data)


_SQLITE_MAGIC = b"SQLite format 3\x00"


@admin_router.get("/backup")
def download_backup():
    if not os.path.exists(DB_PATH):
        raise HTTPException(status_code=404, detail="Database file not found")
    filename = f"clinic_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    return FileResponse(DB_PATH, media_type="application/octet-stream", filename=filename)


@admin_router.post("/restore", status_code=200)
async def restore_backup(file: UploadFile = File(..., description="SQLite .db backup file")):
    data = await file.read()
    if not data.startswith(_SQLITE_MAGIC):
        raise HTTPException(status_code=400, detail="Invalid file: not a SQLite database")
    # Write to a temp file first, then atomically replace
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".db")
    try:
        with os.fdopen(tmp_fd, "wb") as f:
            f.write(data)
        shutil.move(tmp_path, DB_PATH)
    except Exception:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        raise HTTPException(status_code=500, detail="Failed to restore database")
    return {"detail": "Database restored successfully"}


@admin_router.get("/export-excel")
def export_excel(db: Session = Depends(get_db)):
    """Export all records as an Excel file matching the import format (sheet: List)."""
    visits = (
        db.query(OpdVisit)
        .options(
            selectinload(OpdVisit.patients),
            selectinload(OpdVisit.phones),
            selectinload(OpdVisit.owners),
        )
        .order_by(OpdVisit.opd_number)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "List"

    # Header row — matches import column order
    headers = ["OPD#", "Owner Name", "Owner Name 2", "Pet Name", "Pet Type", "Phone"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for visit in visits:
        owners = [o.owner_name for o in visit.owners]
        phones = [p.phone for p in visit.phones]
        owner1 = owners[0] if len(owners) > 0 else None
        owner2 = owners[1] if len(owners) > 1 else None
        phone  = phones[0] if phones else None

        if not visit.patients:
            # OPD with no patients — still export one row
            ws.append([visit.opd_number, owner1, owner2, None, None, phone])
        else:
            for patient in visit.patients:
                ws.append([
                    visit.opd_number,
                    owner1,
                    owner2,
                    patient.pet_name,
                    patient.pet_type,
                    phone,
                ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"clinic_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
